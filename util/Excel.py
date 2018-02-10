# coding:utf-8

from openpyxl import *
from openpyxl.styles import Font
from ProjectVar.var import test_data_excel_path
from FormatTime import date_time_chinese


class ParseExcel(object):
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.workbook = load_workbook(excel_file_path)
        self.font = Font(color=None)
        self.colorDict = {"red": 'FFFF3030', "green": 'FF008B00'}
        self.sheet = self.workbook.active

    # 通过索引（序号）设置需要操作的sheet
    def set_sheet_by_index(self, sheet_index):
        self.sheet = self.get_default_by_index(sheet_index)

    # 通过名称设置需要操作的sheet
    def set_sheet_by_name(self, name):
        self.sheet = self.get_default_by_name(name)

    # 获取当前操作的sheet的title名称
    def get_default_name(self):
        # self.sheet = self.workbook.get_active_sheet
        return self.sheet.title

    # 通过名称获取操作的sheet
    def get_default_by_name(self, name):
        self.sheet = self.workbook.get_sheet_by_name(name)
        return self.sheet

    # 通过序号获取操作的sheet
    def get_default_by_index(self, sheet_index):
        sheets_name = self.workbook.get_sheet_names()
        self.sheet = self.workbook.get_sheet_by_name(sheets_name[sheet_index])
        return self.sheet

    # 获取sheet中最大的行号，从0开始
    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取sheet中最大的列号，从0开始
    def get_max_col_no(self):
        return self.sheet.max_column

    # 获取正在操作sheet中的所有行对象
    def get_all_rows(self):
        # self.sheet.iter_cols
        return list(self.sheet.rows)

    # 获取正在操作sheet中的所有列对象
    def get_all_cols(self):
        ''' cols = []
        for col in self.sheet.iter_rows:
            cols.append(col) '''
        return list(self.sheet.columns)

    # 获取某一个行对象
    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    # 获取某一个列对象
    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    # 获取某一个单元格对象
    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    # 获取某一个单元格的值
    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 给某个单元格写入指定的内容
    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        self.workbook.save(self.excel_file_path)

    # 给某个单元格写入当前时间
    def write_cell_current_time(self, row_no, col_no):
        self.sheet.cell(row=row_no, column=col_no).value = date_time_chinese()
        self.workbook.save(self.excel_file_path)

    # 保存对所有单元格的修改
    def save_excel_file(self):
        self.workbook.save(self.excel_file_path)

if __name__ == '__main__':
    # 测试所有的方法
    print '111'
    excle = ParseExcel(test_data_excel_path)
    sheet1 = excle.get_default_by_name(u'126账号')
    sheet2 = excle.get_default_by_index(0)
    print sheet1, sheet2
    print excle.get_max_row_no(), excle.get_max_col_no()
    # excle.set_sheet_by_index(1)
    excle.set_sheet_by_name(u'联系人')
    print excle.get_default_name()
    rows = excle.get_all_cols()
    print rows
    print excle.get_single_col(1)
    print excle.get_single_row(1)
    excle.write_cell_content(1, 1, u'我就是')
    excle.write_cell_current_time(8, 1)